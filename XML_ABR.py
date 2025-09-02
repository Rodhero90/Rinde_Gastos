import pandas as pd
import xml.etree.ElementTree as ET
import os
from pathlib import Path
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict


class CatalogadorXMLsCFDI:
    """
    Catalogador y analizador de XMLs CFDI
    Genera un catálogo completo con clasificación por categorías corregidas
    """

    def __init__(self, carpetas_cfdi):
        if isinstance(carpetas_cfdi, str):
            self.carpetas_cfdi = [carpetas_cfdi]
        else:
            self.carpetas_cfdi = carpetas_cfdi

        self.namespaces = {
            'cfdi': 'http://www.sat.gob.mx/cfd/4',
            'cfdi3': 'http://www.sat.gob.mx/cfd/3',
            'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital'
        }

        # Categorías corregidas basadas en el archivo Excel
        self.categorias_corregidas = {
            # Materiales de construcción y mantenimiento
            'Materia Prima Construcción/Mantenimiento': {
                'emisores': [
                    'econocomex internacional', 'refrigeracion starr', 'home depot', 'truper',
                    'sherwin williams', 'comex', 'berel', 'elektra', 'liverpool', 'coppel',
                    'cemex', 'holcim', 'cruz azul', 'materiales la union', 'ferreteria',
                    'material para construccion', 'aceros', 'tubos y conexiones'
                ],
                'palabras_clave': [
                    'tuberia', 'tubo', 'valvula', 'codo', 'cople', 'aislante', 'termico',
                    'cemento', 'concreto', 'varilla', 'alambre', 'malla', 'tabique',
                    'pintura', 'thinner', 'brocha', 'rodillo', 'sellador', 'impermeabilizante',
                    'tornillo', 'clavo', 'taquete', 'ancla', 'pija', 'rondana',
                    'cable', 'alambre', 'foco', 'led', 'contacto', 'apagador',
                    'arena', 'grava', 'cal', 'mortero', 'yeso', 'adhesivo',
                    'lija', 'disco', 'broca', 'herramienta', 'martillo', 'desarmador',
                    'material construccion', 'ferreteria', 'plomeria', 'electricidad'
                ]
            },

            # Gasolina y combustibles
            'Gasolina': {
                'emisores': ['pemex', 'bp', 'shell', 'total', 'mobil', 'texaco', 'valdori', 'estacion servicio'],
                'palabras_clave': ['gasolina', 'diesel', 'combustible', 'magna', 'premium', 'litros']
            },

            # Hospedaje
            'Hospedaje': {
                'emisores': ['hotel', 'motel', 'inn', 'resort', 'hostal', 'hostel', 'marriott', 'hilton', 'holiday'],
                'palabras_clave': ['hospedaje', 'habitacion', 'noche', 'estancia', 'alojamiento', 'hotel', 'motel']
            },

            # Transporte
            'Transporte público/privado': {
                'emisores': ['uber', 'didi', 'taxi', 'aeromexico', 'volaris', 'viva aerobus', 'interjet'],
                'palabras_clave': ['transporte', 'viaje', 'pasaje', 'boleto', 'taxi', 'uber', 'avion', 'vuelo']
            },

            # Estacionamiento
            'Estacionamiento': {
                'emisores': ['estacionamiento', 'parking'],
                'palabras_clave': ['estacionamiento', 'parking', 'pension']
            },

            # Arrendamiento
            'Arrendamiento Oficinas': {
                'emisores': [],
                'palabras_clave': ['arrendamiento', 'renta oficina', 'alquiler oficina', 'renta inmueble']
            },

            'Arrendamiento automoviles': {
                'emisores': ['hertz', 'avis', 'budget', 'europcar', 'alamo'],
                'palabras_clave': ['renta auto', 'renta vehiculo', 'alquiler vehiculo', 'car rental']
            },

            # Honorarios
            'Honorarios Contables': {
                'emisores': ['contador', 'contable', 'despacho contable'],
                'palabras_clave': ['honorarios contable', 'servicios contables', 'contabilidad', 'auditoria']
            },

            'Honorarios Legales': {
                'emisores': ['abogado', 'licenciado', 'despacho juridico', 'bufete'],
                'palabras_clave': ['honorarios legales', 'servicios juridicos', 'asesoria legal', 'abogado']
            },

            'Honorarios Administrativos': {
                'emisores': [],
                'palabras_clave': ['honorarios administrativos', 'servicios administrativos', 'gestion administrativa']
            },

            'Honorarios Profesionales': {
                'emisores': [],
                'palabras_clave': ['honorarios profesionales', 'servicios profesionales', 'consultoria']
            },

            'Honorarios Tecnología': {
                'emisores': [],
                'palabras_clave': ['honorarios tecnologia', 'servicios tecnologicos', 'desarrollo software',
                                   'soporte tecnico']
            },

            # Servicios
            'Servicio Administrativo': {
                'emisores': [],
                'palabras_clave': ['servicio administrativo', 'gestion administrativa', 'tramites']
            },

            'Mantenimiento': {
                'emisores': [],
                'palabras_clave': ['mantenimiento', 'reparacion', 'servicio tecnico']
            },

            'Mantenimiento Equipo Transporte': {
                'emisores': [],
                'palabras_clave': ['mantenimiento vehiculo', 'reparacion auto', 'servicio automotriz',
                                   'taller mecanico']
            },

            # Consumo y comida
            'Consumo': {
                'emisores': ['oxxo', 'seven eleven', '7-eleven', 'soriana', 'walmart', 'supermarket'],
                'palabras_clave': ['alimento', 'comida', 'bebida', 'restaurant', 'cafe', 'desayuno', 'comida']
            },

            # Equipo y mobiliario
            'Equipo de Computo': {
                'emisores': ['dell', 'hp', 'lenovo', 'apple', 'microsoft', 'office depot', 'best buy'],
                'palabras_clave': ['computadora', 'laptop', 'monitor', 'teclado', 'mouse', 'impresora',
                                   'equipo computo']
            },

            'Mobiliario y Equipo': {
                'emisores': ['office depot', 'officemax', 'muebles'],
                'palabras_clave': ['mobiliario', 'escritorio', 'silla', 'mesa', 'mueble', 'archivero']
            },

            # Servicios específicos
            'Teléfono': {
                'emisores': ['telcel', 'movistar', 'at&t', 'unefon', 'telmex'],
                'palabras_clave': ['telefono', 'celular', 'plan', 'linea telefonica', 'comunicacion']
            },

            'Licencia Software': {
                'emisores': ['microsoft', 'adobe', 'oracle', 'sap'],
                'palabras_clave': ['licencia software', 'suscripcion software', 'office 365', 'adobe creative']
            },

            # Financieros
            'Comisión Bancaria': {
                'emisores': ['banco', 'banamex', 'bbva', 'santander', 'banorte', 'scotiabank'],
                'palabras_clave': ['comision bancaria', 'comision', 'banco', 'transferencia']
            },

            'Intereses': {
                'emisores': [],
                'palabras_clave': ['intereses', 'interes moratorio', 'rendimiento']
            },

            # Otros
            'Papeleria': {
                'emisores': ['office depot', 'officemax', 'lumen'],
                'palabras_clave': ['papeleria', 'papel', 'folder', 'engrapadora', 'lapiz', 'pluma']
            },

            'Publicidad': {
                'emisores': [],
                'palabras_clave': ['publicidad', 'marketing', 'promocion', 'anuncio']
            },

            'Casetas': {
                'emisores': ['autopistas', 'peaje'],
                'palabras_clave': ['caseta', 'peaje', 'autopista', 'cuota']
            },

            'Derechos': {
                'emisores': [],
                'palabras_clave': ['derechos', 'tramite', 'gobierno']
            },

            'Impuestos': {
                'emisores': [],
                'palabras_clave': ['impuesto', 'predial', 'tenencia', 'sat']
            },

            'Seguro Gastos Médicos': {
                'emisores': ['gnp', 'axa', 'metlife', 'seguros'],
                'palabras_clave': ['seguro medico', 'gastos medicos', 'poliza', 'prima']
            },

            'Servicio Rastreo': {
                'emisores': [],
                'palabras_clave': ['rastreo', 'gps', 'localizacion', 'monitoreo']
            },

            'Promotoria': {
                'emisores': [],
                'palabras_clave': ['promotoria', 'promocion', 'comision venta']
            }
        }

    def clasificar_xml_corregido(self, datos):
        """
        Clasifica el XML usando las categorías corregidas del Excel
        """
        texto_busqueda = f"{datos['emisor_nombre']} {datos['descripcion_concatenada']}".lower()

        # DEBUG: Mostrar algunos ejemplos de clasificación
        debug_ejemplos = ["home depot", "pemex", "hotel", "uber", "comex"]
        if any(ejemplo in texto_busqueda for ejemplo in debug_ejemplos):
            print(f"\nDEBUG Clasificación: {datos['emisor_nombre']}")
            print(f"  Texto búsqueda: {texto_busqueda[:100]}...")

        # Buscar categoría
        mejor_categoria = 'Miscelaneos'
        mejor_puntuacion = 0

        for categoria, criterios in self.categorias_corregidas.items():
            puntuacion = 0

            # Verificar emisores
            for emisor in criterios['emisores']:
                if emisor.lower() in texto_busqueda:
                    puntuacion += 5  # Mayor peso para emisores
                    if any(ejemplo in texto_busqueda for ejemplo in debug_ejemplos):
                        print(f"    Emisor match: {emisor} -> {categoria} (+5)")

            # Verificar palabras clave
            for palabra in criterios['palabras_clave']:
                if palabra.lower() in texto_busqueda:
                    puntuacion += 1
                    if any(ejemplo in texto_busqueda for ejemplo in debug_ejemplos):
                        print(f"    Palabra match: {palabra} -> {categoria} (+1)")

            if puntuacion > mejor_puntuacion:
                mejor_puntuacion = puntuacion
                mejor_categoria = categoria

        # Casos especiales basados en tipo de comprobante
        if datos['tipo_comprobante'] == 'P':  # Pagos
            mejor_categoria = 'N/A'
        elif datos['tipo_comprobante'] == 'N':  # Nómina
            mejor_categoria = 'Honorarios Profesionales'

        datos['categoria'] = mejor_categoria
        datos['confianza_categoria'] = mejor_puntuacion

        if any(ejemplo in texto_busqueda for ejemplo in debug_ejemplos):
            print(f"    RESULTADO: {mejor_categoria} (puntuación: {mejor_puntuacion})")

        # Verificar si es construcción (mantener esta funcionalidad)
        palabras_construccion = [
            'tubo', 'valvula', 'codo', 'cemento', 'adhesivo', 'pintura', 'brocha',
            'cable', 'foco', 'lampara', 'tornillo', 'clavo', 'herramienta', 'lija',
            'silicon', 'pegamento', 'varilla', 'alambre', 'malla', 'tabique',
            'grava', 'arena', 'mortero', 'yeso', 'impermeabilizante', 'sellador'
        ]

        datos['es_construccion'] = any(palabra in texto_busqueda for palabra in palabras_construccion)

        # Extraer palabras clave
        palabras = texto_busqueda.split()
        palabras_relevantes = [p for p in palabras if len(p) > 4 and not p.isdigit()]
        datos['palabras_clave'] = list(set(palabras_relevantes[:10]))

        return datos

    def leer_xml_completo(self, archivo_xml):
        """
        Lee un archivo XML y extrae TODA la información relevante
        """
        try:
            tree = ET.parse(archivo_xml)
            root = tree.getroot()

            # Determinar versión y namespace
            version = root.get('Version', '4.0')
            ns = self.namespaces.copy()
            if version.startswith('3'):
                ns['cfdi'] = ns['cfdi3']

            # Datos básicos del comprobante
            datos = {
                # Identificación del archivo
                'archivo': archivo_xml.name,
                'carpeta': archivo_xml.parent.name,
                'ruta_completa': str(archivo_xml),

                # Datos del comprobante
                'version_cfdi': version,
                'serie': root.get('Serie', ''),
                'folio': root.get('Folio', ''),
                'fecha': root.get('Fecha', ''),
                'fecha_parsed': None,
                'mes': '',
                'año': '',
                'tipo_comprobante': root.get('TipoDeComprobante', ''),
                'tipo_comprobante_desc': self.obtener_tipo_comprobante(root.get('TipoDeComprobante', '')),
                'lugar_expedicion': root.get('LugarExpedicion', ''),
                'metodo_pago': root.get('MetodoPago', ''),
                'metodo_pago_desc': self.obtener_metodo_pago(root.get('MetodoPago', '')),
                'forma_pago': root.get('FormaPago', ''),
                'forma_pago_desc': self.obtener_forma_pago(root.get('FormaPago', '')),
                'condiciones_pago': root.get('CondicionesDePago', ''),
                'moneda': root.get('Moneda', 'MXN'),
                'tipo_cambio': root.get('TipoCambio', '1'),

                # Montos
                'subtotal': float(root.get('SubTotal', '0')),
                'descuento': float(root.get('Descuento', '0')),
                'total': float(root.get('Total', '0')),

                # Emisor
                'emisor_rfc': '',
                'emisor_nombre': '',
                'emisor_regimen': '',

                # Receptor
                'receptor_rfc': '',
                'receptor_nombre': '',
                'receptor_uso_cfdi': '',
                'receptor_uso_cfdi_desc': '',
                'receptor_domicilio_fiscal': '',
                'receptor_regimen': '',

                # Timbre fiscal
                'uuid': '',
                'fecha_timbrado': '',
                'sello_sat': '',
                'no_certificado_sat': '',
                'rfc_prov_certif': '',

                # Conceptos
                'num_conceptos': 0,
                'conceptos': [],
                'descripcion_concatenada': '',
                'claves_productos': [],
                'claves_unidades': [],

                # Impuestos
                'total_impuestos_trasladados': 0,
                'total_impuestos_retenidos': 0,
                'tiene_iva': False,
                'tiene_isr': False,
                'tiene_ieps': False,

                # Clasificación
                'categoria': '',
                'subcategoria': '',
                'es_construccion': False,
                'confianza_categoria': 0,

                # Análisis adicional
                'palabras_clave': [],
                'es_cancelado': False,
                'tiene_addenda': False,
                'tiene_complemento': False
            }

            # Parsear fecha
            if datos['fecha']:
                try:
                    fecha_obj = datetime.strptime(datos['fecha'][:19], "%Y-%m-%dT%H:%M:%S")
                    datos['fecha_parsed'] = fecha_obj
                    datos['mes'] = fecha_obj.strftime("%m-%B")
                    datos['año'] = fecha_obj.year
                except:
                    pass

            # Emisor
            emisor = root.find('.//cfdi:Emisor', ns)
            if emisor is not None:
                datos['emisor_rfc'] = emisor.get('Rfc', '')
                datos['emisor_nombre'] = emisor.get('Nombre', '')
                datos['emisor_regimen'] = emisor.get('RegimenFiscal', '')

            # Receptor
            receptor = root.find('.//cfdi:Receptor', ns)
            if receptor is not None:
                datos['receptor_rfc'] = receptor.get('Rfc', '')
                datos['receptor_nombre'] = receptor.get('Nombre', '')
                datos['receptor_uso_cfdi'] = receptor.get('UsoCFDI', '')
                datos['receptor_uso_cfdi_desc'] = self.obtener_uso_cfdi(receptor.get('UsoCFDI', ''))
                datos['receptor_domicilio_fiscal'] = receptor.get('DomicilioFiscalReceptor', '')
                datos['receptor_regimen'] = receptor.get('RegimenFiscalReceptor', '')

            # Timbre fiscal
            timbre = root.find('.//tfd:TimbreFiscalDigital', ns)
            if timbre is not None:
                datos['uuid'] = timbre.get('UUID', '')
                datos['fecha_timbrado'] = timbre.get('FechaTimbrado', '')
                datos['sello_sat'] = timbre.get('SelloSAT', '')[:50] + '...' if timbre.get('SelloSAT', '') else ''
                datos['no_certificado_sat'] = timbre.get('NoCertificadoSAT', '')
                datos['rfc_prov_certif'] = timbre.get('RfcProvCertif', '')

            # Conceptos
            conceptos = root.findall('.//cfdi:Concepto', ns)
            descripciones = []

            for concepto in conceptos:
                desc = concepto.get('Descripcion', '')
                clave_prod = concepto.get('ClaveProdServ', '')
                clave_unidad = concepto.get('ClaveUnidad', '')
                cantidad = float(concepto.get('Cantidad', '1'))
                unidad = concepto.get('Unidad', '')
                valor_unitario = float(concepto.get('ValorUnitario', '0'))
                importe = float(concepto.get('Importe', '0'))
                descuento = float(concepto.get('Descuento', '0'))

                concepto_dict = {
                    'descripcion': desc,
                    'clave_producto': clave_prod,
                    'clave_unidad': clave_unidad,
                    'cantidad': cantidad,
                    'unidad': unidad,
                    'valor_unitario': valor_unitario,
                    'importe': importe,
                    'descuento': descuento
                }

                datos['conceptos'].append(concepto_dict)

                if desc:
                    descripciones.append(desc)
                    datos['claves_productos'].append(clave_prod)
                    datos['claves_unidades'].append(clave_unidad)

            datos['num_conceptos'] = len(conceptos)
            datos['descripcion_concatenada'] = " | ".join(descripciones[:5])  # Máximo 5 conceptos

            # Impuestos
            impuestos = root.find('.//cfdi:Impuestos', ns)
            if impuestos is not None:
                datos['total_impuestos_trasladados'] = float(impuestos.get('TotalImpuestosTrasladados', '0'))
                datos['total_impuestos_retenidos'] = float(impuestos.get('TotalImpuestosRetenidos', '0'))

                # Verificar tipos de impuestos
                for traslado in root.findall('.//cfdi:Traslado', ns):
                    impuesto = traslado.get('Impuesto', '')
                    if impuesto == '002':
                        datos['tiene_iva'] = True
                    elif impuesto == '003':
                        datos['tiene_ieps'] = True

                for retencion in root.findall('.//cfdi:Retencion', ns):
                    if retencion.get('Impuesto', '') == '001':
                        datos['tiene_isr'] = True

            # Verificar complementos y addendas
            if root.find('.//cfdi:Complemento', ns) is not None:
                datos['tiene_complemento'] = True

            if root.find('.//cfdi:Addenda', ns) is not None:
                datos['tiene_addenda'] = True

            # Clasificar usando el método corregido
            datos = self.clasificar_xml_corregido(datos)

            return datos

        except Exception as e:
            print(f"Error leyendo {archivo_xml}: {str(e)}")
            return None

    def obtener_tipo_comprobante(self, codigo):
        """Devuelve la descripción del tipo de comprobante"""
        tipos = {
            'I': 'Ingreso',
            'E': 'Egreso',
            'T': 'Traslado',
            'N': 'Nómina',
            'P': 'Pago'
        }
        return tipos.get(codigo, codigo)

    def obtener_metodo_pago(self, codigo):
        """Devuelve la descripción del método de pago"""
        metodos = {
            'PUE': 'Pago en una sola exhibición',
            'PPD': 'Pago en parcialidades o diferido'
        }
        return metodos.get(codigo, codigo)

    def obtener_forma_pago(self, codigo):
        """Devuelve la descripción de la forma de pago"""
        formas = {
            '01': 'Efectivo',
            '02': 'Cheque nominativo',
            '03': 'Transferencia electrónica',
            '04': 'Tarjeta de crédito',
            '28': 'Tarjeta de débito',
            '99': 'Por definir'
        }
        return formas.get(codigo, codigo)

    def obtener_uso_cfdi(self, codigo):
        """Devuelve la descripción del uso del CFDI"""
        usos = {
            'G01': 'Adquisición de mercancías',
            'G02': 'Devoluciones, descuentos o bonificaciones',
            'G03': 'Gastos en general',
            'I01': 'Construcciones',
            'I02': 'Mobiliario y equipo de oficina',
            'I03': 'Equipo de transporte',
            'I04': 'Equipo de cómputo',
            'I05': 'Dados, troqueles, moldes',
            'I06': 'Comunicaciones telefónicas',
            'I07': 'Comunicaciones satelitales',
            'I08': 'Otra maquinaria y equipo',
            'D01': 'Honorarios médicos y gastos hospitalarios',
            'D02': 'Gastos médicos por incapacidad',
            'D03': 'Gastos funerales',
            'D04': 'Donativos',
            'D05': 'Intereses hipotecarios',
            'D06': 'Aportaciones voluntarias al SAR',
            'D07': 'Primas por seguros de gastos médicos',
            'D08': 'Gastos de transportación escolar',
            'D09': 'Depósitos en cuentas para el ahorro',
            'D10': 'Pagos por servicios educativos',
            'S01': 'Sin efectos fiscales',
            'CP01': 'Pagos',
            'CN01': 'Nómina'
        }
        return usos.get(codigo, codigo)

    def generar_catalogo_excel(self, archivo_salida):
        """
        Genera un catálogo Excel completo con múltiples hojas usando categorías corregidas
        """
        print("\n📊 Generando catálogo de XMLs con categorías corregidas...")

        # Leer todos los XMLs
        todos_xmls = []

        for carpeta in self.carpetas_cfdi:
            if not os.path.exists(carpeta):
                print(f"⚠️  Carpeta no encontrada: {carpeta}")
                continue

            print(f"\n📁 Procesando: {carpeta}")
            count = 0

            for archivo_xml in Path(carpeta).glob("*.xml"):
                datos = self.leer_xml_completo(archivo_xml)
                if datos:
                    todos_xmls.append(datos)
                    count += 1

                    if count % 50 == 0:
                        print(f"   Procesados: {count} XMLs...")

            print(f"   ✅ Total en {os.path.basename(carpeta)}: {count}")

        print(f"\n📊 Total XMLs procesados: {len(todos_xmls)}")

        if not todos_xmls:
            print("❌ No se encontraron XMLs para procesar")
            return

        # Convertir a DataFrame
        df = pd.DataFrame(todos_xmls)

        # DEBUG: Verificar categorías antes de eliminar duplicados
        print(f"\n🔍 Verificando clasificación inicial...")
        if 'categoria' in df.columns:
            categorias_inicial = df['categoria'].value_counts()
            print(f"   Categorías encontradas: {len(categorias_inicial)}")
            for cat, count in categorias_inicial.head(5).items():
                print(f"   - {cat}: {count}")
        else:
            print("   ❌ ERROR: No se encontró columna 'categoria'")

        # Identificar y eliminar duplicados por UUID
        print("\n🔍 Buscando duplicados...")
        duplicados_antes = len(df)

        # Marcar duplicados antes de eliminarlos
        df['es_duplicado'] = df.duplicated(subset=['uuid'], keep='first')
        num_duplicados = df['es_duplicado'].sum()

        if num_duplicados > 0:
            print(f"⚠️  Encontrados {num_duplicados} XMLs duplicados")
            df_duplicados = df[df['es_duplicado']].copy()
            df = df[~df['es_duplicado']].copy()
            df = df.drop('es_duplicado', axis=1)
            print(f"✅ Duplicados eliminados: {num_duplicados}")
            print(f"📊 XMLs únicos restantes: {len(df)}")
        else:
            print("✅ No se encontraron duplicados")
            df = df.drop('es_duplicado', axis=1)
            df_duplicados = pd.DataFrame()

        # Crear libro de Excel con las mismas hojas que antes pero con categorías corregidas
        wb = Workbook()

        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_font = Font(color="FFFFFF", bold=True)
        highlight_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 1. HOJA RESUMEN
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        ws_resumen.merge_cells('A1:E1')
        ws_resumen['A1'] = "CATÁLOGO DE COMPROBANTES FISCALES (CFDI) - CATEGORÍAS CORREGIDAS"
        ws_resumen['A1'].font = Font(size=16, bold=True)
        ws_resumen['A1'].alignment = Alignment(horizontal='center')

        # Estadísticas
        row = 3
        estadisticas = [
            ["ESTADÍSTICAS GENERALES", "", ""],
            ["Total de XMLs:", len(df), ""],
            ["Periodo:", f"{df['fecha'].min()[:10]} a {df['fecha'].max()[:10]}", ""],
            ["Total facturado:", f"${df['total'].sum():,.2f}", ""],
            ["", "", ""],
            ["POR TIPO DE COMPROBANTE", "", ""],
        ]

        # Agregar conteo por tipo
        for tipo, count in df['tipo_comprobante_desc'].value_counts().items():
            estadisticas.append(
                [f"  {tipo}:", count, f"${df[df['tipo_comprobante_desc'] == tipo]['total'].sum():,.2f}"])

        estadisticas.extend([
            ["", "", ""],
            ["POR CATEGORÍA CORREGIDA", "", ""],
        ])

        # Agregar conteo por categoría corregida
        if 'categoria' in df.columns and len(df) > 0:
            categorias_unicas = df['categoria'].value_counts()
            for cat, count in categorias_unicas.items():
                total_cat = df[df['categoria'] == cat]['total'].sum()
                estadisticas.append([f"  {cat}:", count, f"${total_cat:,.2f}"])
        else:
            estadisticas.append(["  Error: Sin categorías definidas", "", ""])

        estadisticas.extend([
            ["", "", ""],
            ["POR MES", "", ""],
        ])

        # Agregar por mes
        if len(df) > 0:
            df_por_mes = df.groupby(['año', 'mes'])['total'].agg(['count', 'sum']).reset_index()
            for idx in range(len(df_por_mes)):
                año = df_por_mes.iloc[idx]['año']
                mes = df_por_mes.iloc[idx]['mes']
                cantidad = df_por_mes.iloc[idx]['count']
                suma = df_por_mes.iloc[idx]['sum']
                estadisticas.append([f"  {año}-{mes}:", cantidad, f"${suma:,.2f}"])

        # Escribir estadísticas
        for stat in estadisticas:
            ws_resumen[f'A{row}'] = str(stat[0])
            ws_resumen[f'C{row}'] = stat[1] if stat[1] != "" else ""
            ws_resumen[f'E{row}'] = str(stat[2]) if stat[2] != "" else ""

            if stat[0] and isinstance(stat[0], str) and stat[0].isupper() and not stat[0].startswith('  '):
                ws_resumen[f'A{row}'].font = Font(bold=True)
                ws_resumen[f'A{row}'].fill = subheader_fill
                ws_resumen[f'A{row}'].font = subheader_font
            row += 1

        # Ajustar anchos
        ws_resumen.column_dimensions['A'].width = 35
        ws_resumen.column_dimensions['C'].width = 15
        ws_resumen.column_dimensions['E'].width = 20

        # 2. HOJA CATÁLOGO COMPLETO
        ws_catalogo = wb.create_sheet("Catálogo Completo")

        # Seleccionar columnas para mostrar
        columnas_catalogo = [
            'fecha', 'emisor_nombre', 'emisor_rfc', 'total', 'descripcion_concatenada',
            'uuid', 'categoria', 'tipo_comprobante_desc', 'metodo_pago_desc',
            'forma_pago_desc', 'carpeta'
        ]

        df_catalogo = df[columnas_catalogo].copy()
        df_catalogo.columns = [
            'Fecha', 'Emisor', 'RFC Emisor', 'Total', 'Descripción',
            'UUID', 'Categoría', 'Tipo', 'Método Pago', 'Forma Pago', 'Carpeta'
        ]

        # Escribir datos
        for r_idx, row in enumerate(dataframe_to_rows(df_catalogo, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_catalogo.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                # Formato de moneda para columna Total
                if c_idx == 4 and r_idx > 1:  # Columna Total
                    cell.number_format = '$#,##0.00'

                cell.border = border

        # 3. HOJA POR EMISOR
        ws_emisores = wb.create_sheet("Por Emisor")

        # Agrupar por emisor
        df_emisores = df.groupby(['emisor_nombre', 'emisor_rfc', 'categoria']).agg({
            'total': ['count', 'sum'],
            'uuid': 'count'
        }).reset_index()

        df_emisores.columns = ['Emisor', 'RFC', 'Categoría', 'Cantidad', 'Total', 'Facturas']
        df_emisores = df_emisores.sort_values('Total', ascending=False)

        # Escribir datos
        for r_idx, row in enumerate(dataframe_to_rows(df_emisores, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_emisores.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font

                if c_idx == 5 and r_idx > 1:  # Columna Total
                    cell.number_format = '$#,##0.00'

                cell.border = border

        # 4. HOJA CONSTRUCCIÓN
        ws_construccion = wb.create_sheet("Relacionados Construcción")

        df_construccion = df[df['es_construccion'] == True].copy()

        if len(df_construccion) > 0:
            columnas_construccion = [
                'fecha', 'emisor_nombre', 'total', 'descripcion_concatenada',
                'categoria', 'uuid'
            ]

            df_construccion = df_construccion[columnas_construccion]
            df_construccion.columns = [
                'Fecha', 'Emisor', 'Total', 'Descripción', 'Categoría', 'UUID'
            ]

            for r_idx, row in enumerate(dataframe_to_rows(df_construccion, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_construccion.cell(row=r_idx, column=c_idx, value=value)

                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font

                    if c_idx == 3 and r_idx > 1:  # Columna Total
                        cell.number_format = '$#,##0.00'

                    cell.border = border

        # 5. HOJA ANÁLISIS MENSUAL
        ws_mensual = wb.create_sheet("Análisis Mensual")

        # Crear pivot por mes y categoría
        pivot_mensual = pd.pivot_table(
            df,
            values='total',
            index=['año', 'mes'],
            columns='categoria',
            aggfunc='sum',
            fill_value=0
        )

        # Escribir encabezados
        ws_mensual['A1'] = 'Año'
        ws_mensual['B1'] = 'Mes'
        col = 3
        for categoria in pivot_mensual.columns:
            cell = ws_mensual.cell(row=1, column=col, value=categoria)
            cell.fill = header_fill
            cell.font = header_font
            col += 1

        # Escribir datos
        row = 2
        for (año, mes), valores in pivot_mensual.iterrows():
            ws_mensual.cell(row=row, column=1, value=año)
            ws_mensual.cell(row=row, column=2, value=mes)
            col = 3
            for valor in valores:
                cell = ws_mensual.cell(row=row, column=col, value=valor)
                cell.number_format = '$#,##0.00'
                col += 1
            row += 1

        # 6. HOJA TOP GASTOS
        ws_top = wb.create_sheet("Top 50 Gastos")

        df_top = df.nlargest(50, 'total')[
            ['fecha', 'emisor_nombre', 'total', 'descripcion_concatenada', 'categoria', 'uuid']
        ].copy()

        df_top.columns = ['Fecha', 'Emisor', 'Total', 'Descripción', 'Categoría', 'UUID']

        for r_idx, row in enumerate(dataframe_to_rows(df_top, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_top.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font

                if c_idx == 3 and r_idx > 1:  # Columna Total
                    cell.number_format = '$#,##0.00'
                    if float(value) > 10000:  # Resaltar gastos mayores a 10k
                        cell.fill = highlight_fill

                cell.border = border

        # 7. HOJA DE DUPLICADOS (si existen)
        if len(df_duplicados) > 0:
            ws_duplicados = wb.create_sheet("XMLs Duplicados")

            # Título
            ws_duplicados.merge_cells('A1:G1')
            ws_duplicados['A1'] = f"XMLs DUPLICADOS ENCONTRADOS: {len(df_duplicados)} archivos"
            ws_duplicados['A1'].font = Font(size=14, bold=True, color="FF0000")
            ws_duplicados['A1'].alignment = Alignment(horizontal='center')

            # Preparar datos de duplicados
            columnas_dup = ['archivo', 'carpeta', 'uuid', 'fecha', 'emisor_nombre', 'total', 'descripcion_concatenada']
            df_dup_show = df_duplicados[columnas_dup].copy()
            df_dup_show.columns = ['Archivo', 'Carpeta', 'UUID', 'Fecha', 'Emisor', 'Total', 'Descripción']

            # Escribir datos
            for r_idx, row in enumerate(dataframe_to_rows(df_dup_show, index=False, header=True), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_duplicados.cell(row=r_idx, column=c_idx, value=value)

                    if r_idx == 3:  # Headers
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)

                    cell.border = border

                    # Formato para columna Total
                    if c_idx == 6 and r_idx > 3:
                        cell.number_format = '$#,##0.00'

        # Ajustar anchos de columna en todas las hojas
        for ws in wb.worksheets:
            for column_cells in ws.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells if hasattr(cell, 'value'))
                # Obtener la letra de la columna del primer cell no merged
                for cell in column_cells:
                    if hasattr(cell, 'column_letter'):
                        ws.column_dimensions[cell.column_letter].width = min(length + 2, 50)
                        break

        # Guardar archivo
        wb.save(archivo_salida)
        print(f"\n✅ Catálogo con categorías corregidas guardado en: {archivo_salida}")

        # Resumen final
        print("\n📊 RESUMEN DEL CATÁLOGO CORREGIDO:")
        print(f"   - Total XMLs: {len(df)}")
        print(f"   - Total facturado: ${df['total'].sum():,.2f}")
        print(f"   - Emisores únicos: {df['emisor_nombre'].nunique()}")
        print(f"   - Categorías únicas: {df['categoria'].nunique()}")

        print("\n📋 Top 5 categorías por monto:")
        top_categorias = df.groupby('categoria')['total'].sum().sort_values(ascending=False).head()
        for cat, monto in top_categorias.items():
            print(f"   - {cat}: ${monto:,.2f}")


# ========== PROGRAMA PRINCIPAL ==========
if __name__ == "__main__":
    print("\n" + "=" * 80)
    print("📚 CATALOGADOR DE XMLs CFDI - VERSIÓN CORREGIDA")
    print("   • Lee todos los XMLs de las carpetas especificadas")
    print("   • Clasifica usando categorías corregidas del Excel")
    print("   • Genera catálogo Excel con categorías precisas")
    print("=" * 80)

    # Configuración
    carpetas_cfdi = [
        '/Users/gbphy/Downloads/CFDI Junio 2025',
        '/Users/gbphy/Downloads/CFDI Julio 2025'
    ]
    archivo_salida = "/Users/gbphy/Downloads/ABR_JUN_JUL_.xlsx"

    print(f"\n📁 Carpetas a procesar:")
    for carpeta in carpetas_cfdi:
        print(f"   - {carpeta}")
    print(f"📤 Archivo salida: {archivo_salida}")

    # Verificar carpetas
    carpetas_validas = []
    for carpeta in carpetas_cfdi:
        if os.path.exists(carpeta):
            carpetas_validas.append(carpeta)
            print(f"   ✅ {carpeta}")
        else:
            print(f"   ⚠️  No encontrada: {carpeta}")

    if not carpetas_validas:
        print("\n❌ No se encontraron carpetas válidas")
        exit(1)

    # Confirmar
    respuesta = input("\n¿Generar catálogo con categorías corregidas? (s/n): ")

    if respuesta.lower() == 's':
        catalogador = CatalogadorXMLsCFDI(carpetas_validas)

        try:
            catalogador.generar_catalogo_excel(archivo_salida)
            print("\n✅ ¡Catálogo con categorías corregidas generado exitosamente!")

        except Exception as e:
            print(f"\n❌ Error: {e}")
            import traceback

            traceback.print_exc()
    else:
        print("\n❌ Proceso cancelado")